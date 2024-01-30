import openpyxl as xl
import os


class PY_XL:
    def __init__(self, xlPath):
        self.xlPath = xlPath
        self.workbook = None
        self.sheet0 = None
        self.sheet1 = None

    def create_workbook(self):
        self.workbook = xl.Workbook()

    # "Mata Kuliah", "Perkiraan Mahasiswa", "Metode Perkuliahan", "Dosen", "Senin", "Selasa", "Rabu", "Kamis", "Jum'at", "Sabtu", "Maksimal Peserta Per Kelas"
    def create_optimal_timetable_sheet(self):
        self.sheet0 = self.workbook.active
        self.sheet0.title = "JADWAL"
        self.heading0 = [
            "Mata Kuliah", 
            "Perkiraan Mahasiswa", 
            "Metode Perkuliahan",
            "Dosen", 
            "Senin", "Selasa", "Rabu", "Kamis", "Jum'at", "Sabtu", 
            "Maksimal Peserta Per Kelas"
        ]
        self.sheet0.append(self.heading0)

    def create_beban_sks_sheet(self):
        self.workbook.create_sheet("Sheet_A")
        self.sheet1 = self.workbook["Sheet_A"]
        self.sheet1.title = "BEBAN SKS"
        self.heading1 = [
            "Nama", 
            "Mata Kuliah yang diampu", 
            "SKS", 
            "Total SKS",
            "Keterangan"
        ]
        self.sheet1.append(self.heading1)

    def save_workbook(self):
        self.workbook.save(self.xlPath)

    def create_excel_file(self):
        if not os.path.exists(self.xlPath):
            self.create_workbook()
            self.create_optimal_timetable_sheet()
            self.create_beban_sks_sheet()
            self.save_workbook()
        else:
            os.remove(self.xlPath)
            self.create_excel_file()



    # def save_to_excel(self):
    #     xlPath = "assets/datas/omray.xlsx"
    #     if not self.classroom:
    #         classroom = "Regular"
    #     # menyimpan data ke sheet 1
    #     workbook0 = xl.load_workbook(xlPath)
    #     workbook0._active_sheet_index = 0
    #     sheet0 = workbook0.active
    #     sheet0.append([self.waktu, self.selectedSub, classroom])
    #     workbook0.save(xlPath)


